"""MultiSourceParser 单元测试。"""

import tempfile
from pathlib import Path

import pytest

from src.parsing.multi_source import MultiSourceParser


class TestMultiSourceParser:
    """MultiSourceParser 核心行为测试。"""

    def test_empty_directory_returns_empty_status(self, tmp_path):
        parser = MultiSourceParser()
        doc = parser.parse(str(tmp_path))
        assert doc.metadata["input_kind"] == "multi_source"
        assert doc.metadata["parse_status"] == "empty"

    def test_single_file_passthrough(self, tmp_path):
        """单文件输入退化为原注册表派发。"""
        # 创建最小 Excel 文件
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["柜号", "元器件名称", "型号及规格", "数量"])
        ws.append(["K1", "断路器", "NSX250", "2"])
        xlsx = tmp_path / "test.xlsx"
        wb.save(str(xlsx))

        parser = MultiSourceParser()
        doc = parser.parse(str(xlsx))
        assert doc.metadata["input_kind"] == "excel"
        assert len(doc.metadata.get("sheets", [])) >= 1

    def test_directory_merges_multiple_sources(self, tmp_path):
        """目录输入时合并多个解析器的 metadata。"""
        from openpyxl import Workbook

        # Excel
        xlsx = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["柜号", "元器件名称"])
        ws.append(["K1", "断路器"])
        wb.save(str(xlsx))

        # Word（最小 docx — 用 python-docx 如果可用）
        try:
            from docx import Document
            docx_path = tmp_path / "test.docx"
            doc = Document()
            doc.add_paragraph("柜体型采用Blokset，品牌推荐施耐德")
            doc.save(str(docx_path))
        except ImportError:
            docx_path = None

        parser = MultiSourceParser()
        doc = parser.parse(str(tmp_path))

        assert doc.metadata["input_kind"] == "multi_source"
        assert "sheets" in doc.metadata
        if docx_path:
            assert "paragraphs" in doc.metadata

    def test_merge_preserves_source_map(self, tmp_path):
        """合并后 _source_map 记录每个文件的来源格式。"""
        from openpyxl import Workbook

        xlsx = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["柜号", "元器件名称"])
        ws.append(["K1", "断路器"])
        wb.save(str(xlsx))

        parser = MultiSourceParser()
        doc = parser.parse(str(tmp_path))

        source_map = doc.metadata.get("_source_map", {})
        assert len(source_map) >= 1
        assert any("test.xlsx" in k for k in source_map)
