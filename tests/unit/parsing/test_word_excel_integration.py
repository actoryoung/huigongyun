from pathlib import Path

from openpyxl import Workbook

from huigongyun.parsing.excel import ExcelProjectParser
from huigongyun.parsing.registry import build_default_source_registry


def test_excel_multi_sheet_and_records(tmp_path):
    p = Path(tmp_path) / "multi.xlsx"
    wb = Workbook()
    s1 = wb.active
    s1.title = "主元件清单"
    s1.append(["柜号", "物料名称", "规格", "数量", "单位"])
    s1.append(["K1", "断路器", "ModelA", 2, "台"])
    s2 = wb.create_sheet("备注")
    s2.append(["备注"])
    s2.append(["这是测试备注"])
    wb.save(p)

    doc = ExcelProjectParser().parse(str(p))
    assert doc.metadata["input_kind"] == "excel"
    assert doc.metadata["sheet_count"] == 2
    main = next(s for s in doc.metadata["sheets"] if s["name"] == "主元件清单")
    assert main["headers"] == ["柜号", "物料名称", "规格", "数量", "单位"]
    assert main["records"][0]["柜号"] == "K1"
    assert main["records"][0]["数量"] == 2


def test_excel_ignores_empty_headers_and_preserves_row_numbers(tmp_path):
    p = Path(tmp_path) / "empty_headers.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "列表"
    ws.append(["", "物料名称", "数量"])  # first header empty
    ws.append(["", "断路器", 2])
    wb.save(p)

    doc = ExcelProjectParser().parse(str(p))
    rec = doc.metadata["sheets"][0]["records"][0]
    assert "物料名称" in rec
    assert "" not in rec
    assert rec["_row_no"] >= 2


def test_registry_routes_docx_to_word(tmp_path):
    doc = build_default_source_registry().parse(str(Path(tmp_path) / "specification.docx"))
    assert doc.metadata["input_kind"] == "word"
    assert doc.metadata["source_format"] == "word"
