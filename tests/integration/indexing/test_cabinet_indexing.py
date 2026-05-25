from pathlib import Path

from openpyxl import Workbook, load_workbook

from huigongyun.export.spreadsheet import ProjectExporter
from huigongyun.models import ProjectResult
from huigongyun.parsing.excel import ExcelProjectParser
from huigongyun.indexing.cabinets import CabinetIndexBuilder


def test_cabinet_index_builder_returns_markers_for_missing_numbers(tmp_path):
    workbook_path = Path(tmp_path) / "cabinets.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "主元件清单"
    sheet.append(["柜号", "柜型", "物料名称", "数量"])
    sheet.append(["", "进线柜", "断路器", 1])
    sheet.append(["K1", "进线柜", "接触器", 2])
    workbook.save(workbook_path)

    document = ExcelProjectParser().parse(str(workbook_path))
    cabinet_result = CabinetIndexBuilder().build(document)

    assert len(cabinet_result.cabinets) == 2
    assert cabinet_result.cabinets[0].cabinet_no == "UNASSIGNED"
    assert cabinet_result.notes == ["unresolved_cabinet_numbers_present"]
    assert cabinet_result.unresolved_rows[0]["marker"] == "cabinet_no:UNASSIGNED"


def test_cabinet_index_builder_extracts_extended_fields_and_exports_them(tmp_path):
    workbook_path = Path(tmp_path) / "cabinet_fields.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "主元件清单"
    sheet.append(["柜号", "柜型", "外形尺寸", "回路数", "接地方式", "进出线方式", "物料名称", "数量"])
    sheet.append(["K2", "进线柜", "800*2200*600", 12, "TN-S", "上进上出", "断路器", 1])
    workbook.save(workbook_path)

    document = ExcelProjectParser().parse(str(workbook_path))
    cabinet_result = CabinetIndexBuilder().build(document)

    assert len(cabinet_result.cabinets) == 1
    cabinet = cabinet_result.cabinets[0]
    assert cabinet.cabinet_no == "K2"
    assert cabinet.dimensions == "800*2200*600"
    assert cabinet.circuit_count == 12
    assert cabinet.grounding_mode == "TN-S"
    assert cabinet.inbound_outbound == "上进上出"

    export_path = Path(tmp_path) / "export"
    result = ProjectResult(project=document, cabinets=cabinet_result.cabinets)
    exported_paths = ProjectExporter().export(result, str(export_path))

    exported = load_workbook(exported_paths["excel"], data_only=True)
    headers = [cell.value for cell in next(exported["Cabinets"].iter_rows(min_row=1, max_row=1))]
    assert headers == [
        "cabinet_no",
        "cabinet_type",
        "rated_current",
        "dimensions",
        "circuit_count",
        "quantity",
        "inbound_outbound",
        "grounding_mode",
        "confidence",
        "remarks",
    ]
