from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.bootstrap import build_context, build_default_pipeline


def test_pipeline_generates_quote_totals_from_price_sheet(tmp_path):
    workbook_path = Path(tmp_path) / "quote.xlsx"
    workbook = Workbook()

    main_sheet = workbook.active
    main_sheet.title = "主元件清单"
    main_sheet.append(["柜号", "柜型", "物料名称", "规格型号", "单位", "数量", "品牌"])
    main_sheet.append(["K1", "进线柜", "断路器", "MCCB-250A", "台", 2, "施耐德"])

    price_sheet = workbook.create_sheet("价格表")
    price_sheet.append(["物料名称", "规格型号", "品牌", "单价"])
    price_sheet.append(["断路器", "MCCB-250A", "施耐德", 120.5])

    workbook.save(workbook_path)

    result = build_default_pipeline().run(build_context(str(workbook_path), str(tmp_path / "out")))

    # AuxMaterialInjector adds 4 cabinet-type materials for 进线柜 without prices
    assert result.quote_totals["project_total"] == 241.0
    assert result.quote_totals["missing_price_count"] == 4
    assert result.quote_lines[0].subtotal == 241.0
    assert result.quote_lines[0].price_missing is False

    excel_path = Path(result.outputs["excel"])
    exported = load_workbook(excel_path, data_only=True)
    assert "Quote" in exported.sheetnames
    assert "QuoteSummary" in exported.sheetnames


def test_pipeline_marks_missing_quote_prices(tmp_path):
    workbook_path = Path(tmp_path) / "quote_missing.xlsx"
    workbook = Workbook()

    main_sheet = workbook.active
    main_sheet.title = "主元件清单"
    main_sheet.append(["柜号", "柜型", "物料名称", "规格型号", "单位", "数量", "品牌"])
    main_sheet.append(["K1", "进线柜", "断路器", "MCCB-250A", "台", 1, "施耐德"])

    workbook.save(workbook_path)

    result = build_default_pipeline().run(build_context(str(workbook_path), str(tmp_path / "out2")))

    # AuxMaterialInjector adds 4 cabinet-type materials for 进线柜 without prices
    assert result.quote_totals["missing_price_count"] == 5
    assert any(issue.issue_type == "missing_price" for issue in result.issues)
    assert result.quote_lines[0].price_missing is True
