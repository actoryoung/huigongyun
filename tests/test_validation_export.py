from pathlib import Path

from openpyxl import Workbook, load_workbook

from huigongyun.bootstrap import build_context, build_default_pipeline


def test_validation_detects_duplicate_bom_lines_and_export_creates_excel(tmp_path):
    workbook_path = Path(tmp_path) / "dup.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "主元件清单"
    sheet.append(["柜号", "柜型", "物料名称", "规格型号", "单位", "数量", "品牌"])
    sheet.append(["K1", "进线柜", "断路器", "MCCB-250A", "台", 1, "施耐德"])
    sheet.append(["K1", "进线柜", "断路器", "MCCB-250A", "台", 1, "施耐德"])
    workbook.save(workbook_path)

    pipeline = build_default_pipeline()
    result = pipeline.run(build_context(input_path=str(workbook_path), output_dir=str(tmp_path / "out")))

    issue_types = {issue.issue_type for issue in result.issues}
    assert "duplicate_bom_line" in issue_types

    excel_path = Path(result.outputs["excel"])
    assert excel_path.exists()

    exported = load_workbook(excel_path, data_only=True)
    assert "BOM" in exported.sheetnames
    assert "Issues" in exported.sheetnames


def test_validation_marks_unassigned_cabinet_lines(tmp_path):
    workbook_path = Path(tmp_path) / "missing_cabinet.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "主元件清单"
    sheet.append(["柜号", "物料名称", "数量"])
    sheet.append(["", "断路器", 1])
    workbook.save(workbook_path)

    pipeline = build_default_pipeline()
    result = pipeline.run(build_context(input_path=str(workbook_path), output_dir=str(tmp_path / "out2")))

    issue_types = {issue.issue_type for issue in result.issues}
    assert "missing_bom_cabinet_no" in issue_types
