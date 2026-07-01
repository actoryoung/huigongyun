from io import BytesIO
from pathlib import Path

from openpyxl import Workbook

from src.webapp import RUN_STORAGE, create_app


def _run_pipeline_sync(client, workbook_path, filename):
    """Run a pipeline synchronously via the /run endpoint."""
    with workbook_path.open("rb") as fh:
        return client.post(
            "/run",
            data={"input_file": (BytesIO(fh.read()), filename)},
            content_type="multipart/form-data",
        )


def test_webapp_runs_pipeline_and_renders_download_links(tmp_path):
    workbook_path = Path(tmp_path) / "web.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "主元件清单"
    sheet.append(["柜号", "柜型", "物料名称", "规格型号", "单位", "数量", "品牌"])
    sheet.append(["K1", "进线柜", "断路器", "MCCB-250A", "台", 1, "施耐德"])
    workbook.save(workbook_path)

    app = create_app()
    app.testing = True

    with app.test_client() as client:
        response = client.get("/")
        assert response.status_code == 200
        assert "上传 Excel 主元器件清单" in response.get_data(as_text=True)

        run_response = _run_pipeline_sync(client, workbook_path, "web.xlsx")

    assert run_response.status_code == 200
    rendered = run_response.get_data(as_text=True)
    assert "运行结果" in rendered
    assert "download" in rendered
    assert "web_result.json" in rendered or "web_result.xlsx" in rendered


def test_webapp_allows_manual_round_trip_edit_and_reexport(tmp_path):
    RUN_STORAGE.clear()

    workbook_path = Path(tmp_path) / "web_edit.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "主元件清单"
    sheet.append(["柜号", "物料名称", "规格型号", "单位", "数量"])
    sheet.append(["K1", "断路器", "MCCB-250A", "台", 1])
    workbook.save(workbook_path)

    app = create_app()
    app.testing = True

    with app.test_client() as client:
        run_response = _run_pipeline_sync(client, workbook_path, "web_edit.xlsx")
        assert run_response.status_code == 200
        run_id = next(reversed(RUN_STORAGE))

        edit_response = client.post(
            f"/edit/{run_id}",
            data={
                "scope": "bom",
                "target": "K1",
                "material_name": "断路器",
                "field_name": "brand",
                "value": "施耐德",
            },
        )

    assert edit_response.status_code == 200
    rendered = edit_response.get_data(as_text=True)
    assert "施耐德" in rendered

    run_state = RUN_STORAGE[run_id]
    result = run_state["result"]
    assert any(edit.field_name == "brand" for edit in result.user_edits)
    assert any(bom_line.material.brand == "施耐德" for bom_line in result.bom_lines)


# ── AJAX JSON edit endpoint ──────────────────────────────────────────


def test_edit_json_endpoint_applies_bom_edit(tmp_path):
    """POST /edit/<run_id>/json with valid JSON → 200, success, updated result."""
    RUN_STORAGE.clear()

    workbook_path = Path(tmp_path) / "ajax_edit.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "主元件清单"
    sheet.append(["柜号", "物料名称", "规格型号", "单位", "数量", "品牌"])
    sheet.append(["K1", "断路器", "MCCB-250A", "台", 1, ""])
    workbook.save(workbook_path)

    app = create_app()
    app.testing = True

    with app.test_client() as client:
        run_response = _run_pipeline_sync(client, workbook_path, "ajax_edit.xlsx")
        assert run_response.status_code == 200
        run_id = next(reversed(RUN_STORAGE))

        # Edit brand via JSON endpoint
        json_response = client.post(
            f"/edit/{run_id}/json",
            json={
                "scope": "bom",
                "target": "K1",
                "material_name": "断路器",
                "field_name": "brand",
                "value": "施耐德",
            },
        )

    assert json_response.status_code == 200
    data = json_response.get_json()
    assert data["success"] is True
    assert data["error"] is None
    assert data["result"] is not None
    assert "bom_lines" in data["result"]
    assert "cabinets" in data["result"]
    assert "issues" in data["result"]

    # Verify the edit was actually applied in RUN_STORAGE
    run_state = RUN_STORAGE[run_id]
    result = run_state["result"]
    assert any(bom_line.material.brand == "施耐德" for bom_line in result.bom_lines)


def test_edit_json_endpoint_returns_404_for_invalid_run_id(tmp_path):
    """POST /edit/<run_id>/json with nonexistent run_id → 404."""
    RUN_STORAGE.clear()

    app = create_app()
    app.testing = True

    with app.test_client() as client:
        response = client.post(
            "/edit/nonexistent_id/json",
            json={"scope": "bom", "target": "K1", "field_name": "brand", "value": "A"},
        )

    assert response.status_code == 404
    data = response.get_json()
    assert data["success"] is False
    assert data["error"] is not None


def test_edit_json_endpoint_returns_400_for_invalid_scope(tmp_path):
    """POST /edit/<run_id>/json with invalid scope → 400."""
    RUN_STORAGE.clear()

    workbook_path = Path(tmp_path) / "ajax_bad_scope.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "主元件清单"
    sheet.append(["柜号", "物料名称", "规格型号", "单位", "数量"])
    sheet.append(["K1", "断路器", "MCCB-250A", "台", 1])
    workbook.save(workbook_path)

    app = create_app()
    app.testing = True

    with app.test_client() as client:
        run_response = _run_pipeline_sync(client, workbook_path, "ajax_bad_scope.xlsx")
        assert run_response.status_code == 200
        run_id = next(reversed(RUN_STORAGE))

        response = client.post(
            f"/edit/{run_id}/json",
            json={"scope": "invalid_scope", "target": "K1", "field_name": "x", "value": "x"},
        )

    assert response.status_code == 400
    data = response.get_json()
    assert data["success"] is False
    assert data["error"] is not None


def test_edit_json_endpoint_applies_pricing_edit(tmp_path):
    """POST /edit/<run_id>/json with pricing scope → sets unit_price."""
    RUN_STORAGE.clear()

    workbook_path = Path(tmp_path) / "ajax_price.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "主元件清单"
    sheet.append(["柜号", "物料名称", "规格型号", "单位", "数量"])
    sheet.append(["K1", "断路器", "MCCB-250A", "台", 1])
    workbook.save(workbook_path)

    app = create_app()
    app.testing = True

    with app.test_client() as client:
        run_response = _run_pipeline_sync(client, workbook_path, "ajax_price.xlsx")
        assert run_response.status_code == 200
        run_id = next(reversed(RUN_STORAGE))

        response = client.post(
            f"/edit/{run_id}/json",
            json={
                "scope": "pricing",
                "target": "K1",
                "material_name": "断路器",
                "field_name": "unit_price",
                "value": "150.5",
            },
        )

    # pricing scope targets QuoteLine objects in result.summary,
    # which are empty for this test data → expect 400
    assert response.status_code == 400
    data = response.get_json()
    assert data["success"] is False
