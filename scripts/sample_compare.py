#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Run default pipeline on example samples and compare to reference Excel outputs.

Writes per-project reports to `output/sample_comparison/<project>` and an
overall `output/sample_comparison/overall_comparison.json`.

Usage: python scripts/sample_compare.py
"""
from __future__ import annotations

import json
import traceback
from pathlib import Path

from huigongyun.bootstrap import build_default_pipeline, build_context
from openpyxl import load_workbook


OUT_DIR = Path("output/sample_comparison")
EXAMPLE_DIR = Path("example0516.")


def find_reference_excel(project_dir: Path) -> Path | None:
    ref_dir = project_dir / "参考输出_报价结果"
    if not ref_dir.exists():
        return None
    for p in ref_dir.iterdir():
        if p.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
            return p
    return None


def parse_reference_excel(ref_path: Path):
    try:
        wb = load_workbook(ref_path, data_only=True, read_only=True)
    except Exception as e:
        return {"error": f"failed to open workbook: {e}"}

    sheet_names = wb.sheetnames
    preferred = None
    for name in ("Summary", "BOM", "汇总", "报价", "摘要"):
        if name in sheet_names:
            preferred = name
            break
    if not preferred:
        preferred = sheet_names[0] if sheet_names else None
    if not preferred:
        return {"error": "no sheets"}

    sheet = wb[preferred]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {"error": "empty sheet"}
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]

    name_idx = None
    qty_idx = None
    spec_idx = None
    for i, h in enumerate(headers):
        lh = h.lower()
        if any(k in lh for k in ("物料", "名称", "material", "name", "品名")):
            name_idx = i
        if any(k in lh for k in ("数量", "qty", "quantity", "数量(台)")):
            qty_idx = i
        if any(k in lh for k in ("规格", "型号", "spec", "规格型号")):
            spec_idx = i

    items = []
    for r in rows[1:]:
        if name_idx is None:
            continue
        name = r[name_idx] if name_idx < len(r) else None
        if not name:
            continue
        spec = r[spec_idx] if spec_idx is not None and spec_idx < len(r) else None
        qty = r[qty_idx] if qty_idx is not None and qty_idx < len(r) else None
        try:
            qty = float(qty) if qty not in (None, "") else None
        except Exception:
            qty = None
        items.append({"name": str(name).strip(), "spec": str(spec).strip() if spec else None, "quantity": qty})
    return {"sheet": preferred, "headers": headers, "items": items}


def load_generated_json(gen_json_path: Path):
    try:
        return json.loads(gen_json_path.read_text(encoding="utf-8"))
    except Exception as e:
        return {"error": str(e)}


def compare(reference, generated):
    ref_items = reference.get("items", [])
    gen_summary = generated.get("summary", []) or []

    def normalize(n):
        return str(n).strip().lower() if n else ""

    ref_map = {normalize(i["name"]): i for i in ref_items}
    gen_map = {normalize(i.get("name")): i for i in gen_summary}

    missing = []
    extra = []
    qty_mismatch = []
    for name, r in ref_map.items():
        g = gen_map.get(name)
        if g is None:
            missing.append(r)
        else:
            r_qty = r.get("quantity") or 0
            g_qty = g.get("quantity") or 0
            if r_qty != g_qty:
                qty_mismatch.append({"name": r.get("name"), "ref_qty": r_qty, "gen_qty": g_qty})
    for name, g in gen_map.items():
        if name not in ref_map:
            extra.append(g)
    return {"missing": missing, "extra": extra, "qty_mismatch": qty_mismatch}


def run():
    pipeline = build_default_pipeline()
    out_dir = OUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    root = EXAMPLE_DIR
    projects = [p for p in root.iterdir() if p.is_dir()]
    overall = {}
    for proj in projects:
        proj_name = proj.name
        proj_out = out_dir / proj_name
        proj_out.mkdir(parents=True, exist_ok=True)
        inputs_dir = proj / "输入资料"
        if not inputs_dir.exists():
            overall[proj_name] = {"error": "no inputs folder"}
            continue
        inputs = sorted([f for f in inputs_dir.iterdir() if f.is_file()])
        results = {}
        for f in inputs:
            try:
                ctx = build_context(str(f), str(proj_out))
                result = pipeline.run(ctx)
                results[f.name] = {"project": result.project.project_name, "outputs": result.outputs, "project_metadata": result.project.metadata}
            except Exception as e:
                results[f.name] = {"error": str(e), "trace": traceback.format_exc()}

        ref = find_reference_excel(proj)
        if ref:
            parsed_ref = parse_reference_excel(ref)
            gen_jsons = list(proj_out.glob("*_result.json"))
            gen_compare = {}
            if gen_jsons:
                gen = load_generated_json(gen_jsons[0])
                gen_compare = compare(parsed_ref, gen)
            else:
                gen_compare = {"error": "no generated json found"}
            overall[proj_name] = {"inputs_results": results, "reference": str(ref), "reference_parsed": parsed_ref, "comparison": gen_compare}
            (proj_out / f"{proj_name}_comparison.json").write_text(json.dumps(overall[proj_name], ensure_ascii=False, indent=2), encoding="utf-8")
        else:
            overall[proj_name] = {"inputs_results": results, "error": "no reference found"}

    (out_dir / "overall_comparison.json").write_text(json.dumps(overall, ensure_ascii=False, indent=2), encoding="utf-8")
    print("Done. reports written to", out_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(run())
