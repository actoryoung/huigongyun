"""Excel 解析模块：将工作簿转换为规范化的中间表示。

本模块负责发现工作簿并抽取表格级结构，生成供后续 BOM/机柜提取使用的
`sheets` 中间表示。
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..models import ProjectDocument


@dataclass(slots=True)
class SheetSnapshot:
    name: str
    row_count: int
    data_row_count: int
    column_count: int
    headers: list[str]
    sample_rows: list[list[Any]]
    records: list[dict[str, Any]]
    pre_header_meta: dict[str, Any] = field(default_factory=dict)


class ExcelProjectParser:
    """将基于 Excel 的项目输入解析为规范化的文档模型。

    MVP 版本侧重于工作簿发现与表结构抽取，后续阶段可将行提升为机柜、
    BOM 行与校验规则。
    """

    def parse(self, input_path: str) -> ProjectDocument:
        path = Path(input_path)
        source = self._resolve_source(path)
        if source is None:
            return ProjectDocument(
                project_name=path.stem or "project",
                files=[str(path)],
                metadata={
                    "input_kind": "unknown",
                    "parse_status": "missing_input",
                    "message": "Input file not found; returned scaffold document.",
                },
            )

        if source.is_dir():
            return ProjectDocument(
                project_name=path.stem or "project",
                files=[str(source)],
                metadata={
                    "input_kind": "directory",
                    "parse_status": "ambiguous",
                    "candidate_files": self._list_excel_candidates(source),
                    "message": "Directory contains one or more Excel files; choose a specific workbook before parsing.",
                },
            )

        if source.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            return ProjectDocument(
                project_name=source.stem,
                files=[str(source)],
                metadata={
                    "input_kind": "non_excel",
                    "parse_status": "skipped",
                    "message": "Only Excel parsing is implemented in this stage.",
                },
            )

        workbook = load_workbook(source, data_only=True, read_only=True)
        sheets = [self._snapshot_sheet(sheet_name, workbook[sheet_name]) for sheet_name in workbook.sheetnames]

        return ProjectDocument(
            project_name=source.stem,
            files=[str(source)],
            metadata={
                "input_kind": "excel",
                "parse_status": "ok",
                "sheet_count": len(sheets),
                "sheets": [self._sheet_snapshot_to_dict(snapshot) for snapshot in sheets],
            },
        )

    def _resolve_source(self, path: Path) -> Path | None:
        if path.is_file():
            return path
        if path.is_dir():
            if self._list_excel_candidates(path):
                return path
        return None

    def _list_excel_candidates(self, path: Path) -> list[str]:
        return [
            str(candidate)
            for candidate in sorted(path.iterdir())
            if candidate.is_file() and candidate.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}
        ]

    # Keywords that indicate a data header row (not a metadata row)
    _HEADER_KEYWORDS = {
        "柜号", "元器件名称", "物料名称", "元件名称", "设备名称", "品名", "名称",
        "型号及规格", "规格型号", "型号", "规格", "spec",
        "单位", "unit",
        "数量", "qty",
        "品牌", "厂家", "生产厂家", "manufacturer",
        "长交期", "lead_time",
        "设备位号", "功率", "额定电压", "额定电流", "控制方式",
    }
    # Keywords that appear in cabinet-level metadata rows
    _CABINET_META_KEYS = {
        "柜型": "cabinet_type",
        "设备编号": "device_id",
        "外形尺寸": "dimensions",
        "数量": "quantity",
    }
    # Values that indicate a row is a repeated metadata/header section, not data
    _META_NOISE_VALUES = {"设备编号", "外形尺寸", "柜号", "柜型", "数量", "单位", "元器件名称", "型号及规格", "生产厂家", "长交期"}

    def _find_header_row(self, cleaned_rows: list[tuple[int, list[Any]]]) -> int:
        """Find the header row index by keyword density scoring.

        Returns the index (into cleaned_rows) of the best-matching header row,
        or 0 if no clear header is found.
        """
        best_idx = 0
        best_score = 0
        for idx, (_row_no, row) in enumerate(cleaned_rows):
            row_texts = {str(cell).strip() for cell in row if cell not in (None, "")}
            score = len(row_texts & self._HEADER_KEYWORDS)
            if score > best_score:
                best_score = score
                best_idx = idx
        # Require at least 2 keyword matches to consider it a header row
        return best_idx if best_score >= 2 else 0

    def _filter_section_meta_rows(
        self,
        data_rows: list[tuple[int, list[Any]]],
        headers: list[str],
    ) -> list[tuple[int, list[Any]]]:
        """Filter out repeated section metadata/header rows from the data.

        In real-world Excel sheets, each cabinet section may repeat metadata rows
        (设备编号, 外形尺寸) and even the header row (柜号, 元器件名称...).
        These should be excluded from data records.
        """
        # Find the cabinet-no column index (first header that matches CABINET_KEYS)
        cab_col_idx = -1
        cab_keys = {"柜号", "cabinet_no", "柜位", "柜体", "柜名"}
        for i, h in enumerate(headers):
            if h.strip() in cab_keys:
                cab_col_idx = i
                break
        if cab_col_idx < 0:
            return data_rows

        filtered: list[tuple[int, list[Any]]] = []
        for row_no, row in data_rows:
            val = str(row[cab_col_idx]).strip() if cab_col_idx < len(row) else ""
            if val in self._META_NOISE_VALUES:
                continue
            # Also skip rows where all values match header names (exact header repeat)
            non_empty = [str(c).strip() for c in row if c not in (None, "")]
            if len(non_empty) >= 2 and all(
                h in self._HEADER_KEYWORDS or h in self._META_NOISE_VALUES
                for h in non_empty
            ):
                # Check if at least 50% of non-empty cells are known header/metadata terms
                matched = sum(
                    1 for h in non_empty
                    if h in self._HEADER_KEYWORDS or h in self._META_NOISE_VALUES
                )
                if matched >= len(non_empty) * 0.5:
                    continue
            filtered.append((row_no, row))

        return filtered

    def _extract_pre_header_meta(self, cleaned_rows: list[tuple[int, list[Any]]], header_idx: int) -> dict[str, Any]:
        """Extract cabinet-level metadata from rows before the header row.

        Scans metadata rows for key-value pairs like '柜型 BlokSeT' or
        '外形尺寸 1100mm*1000mm*2200mm' and returns structured metadata.
        """
        meta: dict[str, Any] = {}
        for _row_no, row in cleaned_rows[:header_idx]:
            texts = [str(cell).strip() if cell not in (None, "") else "" for cell in row]
            for i, cell_text in enumerate(texts):
                if not cell_text:
                    continue
                cell_clean = cell_text.replace("：", ":").strip()
                for meta_key, meta_field in self._CABINET_META_KEYS.items():
                    if cell_clean == meta_key or cell_clean.startswith(meta_key + ":"):
                        # Look for value in adjacent cells
                        value = self._find_adjacent_value(texts, i)
                        if value:
                            if meta_field not in meta:
                                meta[meta_field] = value
                            else:
                                # Multiple values for same key → list
                                existing = meta[meta_field]
                                if isinstance(existing, list):
                                    existing.append(value)
                                else:
                                    meta[meta_field] = [existing, value]
        return meta

    @staticmethod
    def _find_adjacent_value(texts: list[str], idx: int) -> str | None:
        """Find a value adjacent to a key cell (same row, same or next column)."""
        # Check current cell (key:value in one cell)
        cell_text = texts[idx]
        for sep in (":", "：", " "):
            parts = cell_text.split(sep, 1)
            if len(parts) == 2 and parts[1].strip():
                return parts[1].strip()

        # Check next cells
        for offset in range(1, min(4, len(texts) - idx)):
            val = texts[idx + offset]
            if val and val.strip():
                return val.strip()
        return None

    def _snapshot_sheet(self, sheet_name: str, worksheet: Any) -> SheetSnapshot:
        rows = list(worksheet.iter_rows(values_only=True))
        cleaned_rows: list[tuple[int, list[Any]]] = []
        for row_no, row in enumerate(rows, start=1):
            cleaned_row = self._clean_row(row)
            if any(cell is not None and cell != "" for cell in cleaned_row):
                cleaned_rows.append((row_no, cleaned_row))
        if cleaned_rows:
            # Detect header row instead of assuming row 0
            header_idx = self._find_header_row(cleaned_rows)
            header_row_no, header_row = cleaned_rows[header_idx]
            headers = [str(value).strip() if value is not None else "" for value in header_row]
            # Rows after header are data, rows before are metadata
            raw_data_rows = cleaned_rows[header_idx + 1:]
            # Filter out repeated section metadata/header rows within data
            data_rows = self._filter_section_meta_rows(raw_data_rows, headers)
            sample_rows = data_rows[:5]
            records = [self._row_to_record(headers, row, row_no, sheet_name) for row_no, row in data_rows]
            pre_header_meta = self._extract_pre_header_meta(cleaned_rows, header_idx)
        else:
            headers = []
            sample_rows = []
            records = []
            header_row_no = 0
            pre_header_meta = {}
        column_count = max((len(row) for _, row in cleaned_rows), default=0)
        return SheetSnapshot(
            name=sheet_name,
            row_count=len(cleaned_rows),
            data_row_count=len(records),
            column_count=column_count,
            headers=headers,
            sample_rows=[row for _, row in sample_rows],
            records=records,
            pre_header_meta=pre_header_meta,
        )

    def _clean_row(self, row: tuple[Any, ...]) -> list[Any]:
        return [value if value is not None else "" for value in row]

    def _sheet_snapshot_to_dict(self, snapshot: SheetSnapshot) -> dict[str, Any]:
        return {
            "name": snapshot.name,
            "row_count": snapshot.row_count,
            "data_row_count": snapshot.data_row_count,
            "column_count": snapshot.column_count,
            "headers": snapshot.headers,
            "sample_rows": snapshot.sample_rows,
            "records": snapshot.records,
            "pre_header_meta": snapshot.pre_header_meta,
        }

    def _row_to_record(self, headers: list[str], row: list[Any], row_no: int, sheet_name: str) -> dict[str, Any]:
        record: dict[str, Any] = {"_sheet_name": sheet_name, "_row_no": row_no}
        for index, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[index] if index < len(row) else ""
        return record
