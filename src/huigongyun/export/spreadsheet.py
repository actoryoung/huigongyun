"""将流水线结果导出为本地文件，并可选地上传到 MinIO。

当 MinIO 可用时，导出器使用的环境变量：
    - `MINIO_ENDPOINT`、`MINIO_ACCESS_KEY`、`MINIO_SECRET_KEY`、`MINIO_BUCKET`
    - `MINIO_SECURE`（设置为 "1" 或 "true" 使用 https）
    - `MINIO_PUBLIC_URL`（可选：用于重写 presigned URL 的主机可访问基准 URL）

导出器会先将 JSON 与 Excel 工作簿写入本地，然后尝试上传到配置的
MinIO 桶，并在可能时将 `outputs` 映射替换为 presigned URL。
"""

from __future__ import annotations

import json
import os
from dataclasses import asdict
from datetime import timedelta
from pathlib import Path

from openpyxl import Workbook

from ..models import ProjectResult

try:
    from minio import Minio  # type: ignore
    _HAS_MINIO = True
except Exception:
    _HAS_MINIO = False


class ProjectExporter:
    """导出流水线结果为本地工件，并可选上传到对象存储（MinIO）。

    主要职责：
      1. 将完整的 `ProjectResult` 序列化为本地 JSON 文件并写入 Excel 工作簿；
      2. 当 MinIO 可用时，上传工件并返回对宿主可访问的 presigned URL（可由
         `MINIO_PUBLIC_URL` 覆盖返回 URL 的主机/协议部分）；
      3. 在上传失败时保持对本地路径的回退，以便用户可通过 `download` 端点获取。
    """

    def export(self, result: ProjectResult, output_dir: str) -> dict[str, str]:
        """写入本地 JSON/Excel 并在可能时上传到 MinIO。

        返回一个 `outputs` 字典，键为工件类型（例如 'json'/'excel'），值为本地路径
        或经重写后的 presigned URL。

        注意事项：
          - 函数保证 `output_dir` 已存在；
          - 上传或签名过程中出现异常不会中止函数，而是回退到本地路径；
          - 使用 `MINIO_PUBLIC_URL` 可将 presigned URL 的 host/scheme 替换为宿主可访问的基准。
        """
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        json_path = output_path / f"{result.project.project_name}_result.json"
        excel_path = output_path / f"{result.project.project_name}_result.xlsx"

        # Prepare local artifact paths and attach them to result before writing files.
        outputs: dict[str, str] = {"json": str(json_path), "excel": str(excel_path)}
        result.outputs = outputs

        # write local artifacts (excel first to ensure the file exists if needed)
        self._write_excel(result, excel_path)
        json_path.write_text(json.dumps(asdict(result), ensure_ascii=False, indent=2), encoding="utf-8")

        # If MinIO is configured and client installed, upload artifacts and return presigned URLs
        minio_endpoint = os.environ.get("MINIO_ENDPOINT")
        minio_access_key = os.environ.get("MINIO_ACCESS_KEY")
        minio_secret_key = os.environ.get("MINIO_SECRET_KEY")
        minio_bucket = os.environ.get("MINIO_BUCKET", "huigongyun")
        minio_secure = os.environ.get("MINIO_SECURE", "0").lower() in {"1", "true", "yes"}

        if _HAS_MINIO and minio_endpoint and minio_access_key and minio_secret_key:
            client = Minio(minio_endpoint, access_key=minio_access_key, secret_key=minio_secret_key, secure=minio_secure)
            try:
                if not client.bucket_exists(minio_bucket):
                    client.make_bucket(minio_bucket)
            except Exception:
                pass

            # upload and presign
            for key, path in list(outputs.items()):
                p = Path(path)
                object_name = f"{result.project.project_name}/{p.name}"
                try:
                    client.fput_object(minio_bucket, object_name, str(p))
                    url = client.presigned_get_object(minio_bucket, object_name, expires=timedelta(hours=24))
                    # If a public URL is provided (e.g. http://localhost:9000), rewrite
                    # the presigned URL's scheme/netloc so it is reachable from the host.
                    public_base = os.environ.get("MINIO_PUBLIC_URL")
                    if public_base and isinstance(url, str) and url.startswith("http"):
                        from urllib.parse import urlparse, urlunparse

                        pub = urlparse(public_base)
                        parsed = urlparse(url)
                        new_parsed = parsed._replace(scheme=pub.scheme, netloc=pub.netloc)
                        url = urlunparse(new_parsed)

                    outputs[key] = url
                except Exception:
                    # if upload fails, keep local path
                    outputs[key] = str(p)

        result.outputs = outputs
        return result.outputs

    def _write_excel(self, result: ProjectResult, excel_path: Path) -> None:
        """将 `ProjectResult` 按工作表写入 Excel 文件。

        工作表顺序与内容：Project, Cabinets, BOM, Summary, Quote, QuoteSummary, Issues。
        如果存在默认的空 `Sheet`，则删除后保存文件。
        该函数不抛出上传异常；仅负责本地写入。
        """
        workbook = Workbook()

        self._write_project_sheet(workbook, result)
        self._write_cabinets_sheet(workbook, result)
        self._write_bom_sheet(workbook, result)
        self._write_summary_sheet(workbook, result)
        self._write_quote_sheet(workbook, result)
        self._write_quote_summary_sheet(workbook, result)
        self._write_issues_sheet(workbook, result)

        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])
        workbook.save(excel_path)

    def _write_project_sheet(self, workbook: Workbook, result: ProjectResult) -> None:
        """写入项目元信息（名称、文件清单与 metadata）。"""
        sheet = workbook.create_sheet("Project")
        sheet.append(["project_name", result.project.project_name])
        sheet.append(["files", json.dumps(result.project.files, ensure_ascii=False)])
        sheet.append(["metadata", json.dumps(result.project.metadata, ensure_ascii=False, default=str)])

    def _write_cabinets_sheet(self, workbook: Workbook, result: ProjectResult) -> None:
        """写入机柜信息表，列出每个机柜的核心字段与置信度供人工核对。"""
        sheet = workbook.create_sheet("Cabinets")
        sheet.append([
            "cabinet_no",
            "cabinet_type",
            "rated_current",
            "dimensions",
            "circuit_count",
            "quantity",
            "inbound_outbound",
            "grounding_mode",
            "confidence",
            "remarks",
        ])
        for cabinet in result.cabinets:
            sheet.append([
                cabinet.cabinet_no,
                cabinet.cabinet_type,
                cabinet.rated_current,
                cabinet.dimensions,
                cabinet.circuit_count,
                cabinet.quantity,
                cabinet.inbound_outbound,
                cabinet.grounding_mode,
                cabinet.confidence,
                cabinet.remarks,
            ])

    def _write_bom_sheet(self, workbook: Workbook, result: ProjectResult) -> None:
        """写入 BOM 行表；每行对应一个 `BomLine` 的序列化表示。"""
        sheet = workbook.create_sheet("BOM")
        sheet.append([
            "cabinet_no",
            "material_name",
            "spec",
            "unit",
            "quantity",
            "brand",
            "normalized_name",
            "normalized_spec",
            "confidence",
            "derived_from",
            "risk_tags",
        ])
        for bom_line in result.bom_lines:
            material = bom_line.material
            sheet.append([
                bom_line.cabinet_no,
                material.name,
                material.spec,
                material.unit,
                material.quantity,
                material.brand,
                material.normalized_name,
                material.normalized_spec,
                material.confidence,
                bom_line.derived_from,
                json.dumps(bom_line.risk_tags, ensure_ascii=False),
            ])

    def _write_summary_sheet(self, workbook: Workbook, result: ProjectResult) -> None:
        """写入汇总（聚合后的物料清单）。"""
        sheet = workbook.create_sheet("Summary")
        sheet.append(["material_name", "spec", "unit", "quantity", "brand", "normalized_name", "normalized_spec", "confidence"]) 
        for material in result.summary:
            sheet.append([
                material.name,
                material.spec,
                material.unit,
                material.quantity,
                material.brand,
                material.normalized_name,
                material.normalized_spec,
                material.confidence,
            ])

    def _write_issues_sheet(self, workbook: Workbook, result: ProjectResult) -> None:
        """写入校验/待确认问题列表，便于人工审阅与追溯。"""
        sheet = workbook.create_sheet("Issues")
        sheet.append(["issue_type", "severity", "message", "cabinet_no", "material_name", "details"])
        for issue in result.issues:
            sheet.append([
                issue.issue_type,
                issue.severity,
                issue.message,
                issue.cabinet_no,
                issue.material_name,
                json.dumps(issue.details, ensure_ascii=False, default=str),
            ])

    def _write_quote_sheet(self, workbook: Workbook, result: ProjectResult) -> None:
        """写入逐行报价信息（与 `quote_lines` 对应）。"""
        sheet = workbook.create_sheet("Quote")
        sheet.append([
            "cabinet_no",
            "material_name",
            "spec",
            "unit",
            "quantity",
            "brand",
            "unit_price",
            "subtotal",
            "price_source",
            "price_confidence",
            "price_missing",
            "remarks",
        ])
        for quote_line in result.quote_lines:
            sheet.append([
                quote_line.cabinet_no,
                quote_line.material_name,
                quote_line.spec,
                quote_line.unit,
                quote_line.quantity,
                quote_line.brand,
                quote_line.unit_price,
                quote_line.subtotal,
                quote_line.price_source,
                quote_line.price_confidence,
                quote_line.price_missing,
                quote_line.remarks,
            ])

    def _write_quote_summary_sheet(self, workbook: Workbook, result: ProjectResult) -> None:
        """写入报价汇总表：包含总体指标与按机柜的汇总小计。

        结构：两列 `metric` / `value`，当遇到 `cabinet_totals` 时展开每个机柜的小计。
        """
        sheet = workbook.create_sheet("QuoteSummary")
        sheet.append(["metric", "value"])
        for key, value in result.quote_totals.items():
            if key == "cabinet_totals" and isinstance(value, dict):
                for cabinet_no, subtotal in value.items():
                    sheet.append([f"cabinet_total:{cabinet_no}", subtotal])
            else:
                sheet.append([key, value])
